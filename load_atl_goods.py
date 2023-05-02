import xml.etree.ElementTree as ET
from datetime import time

import openpyxl
import urllib.request

tree = ET.parse('price/atlantmarket.xml')
root = tree.getroot()

data = {}

for offer in root.findall('./shop/offers/offer'):

    vendor_code = offer.find('barcode').text
    offer_data = {}

    if offer.get('available') == "true":
        offer_data['available'] = "В наявності"
    elif offer.get('available') == "false":
        offer_data['available'] = 'Немає в наявності'

    offer_data['price'] = float(offer.find('price').text)
    offer_data['name'] = offer.find('name').text

    if offer.find('categoryId') is not None:
        offer_data['categoryId'] = int(offer.find('categoryId').text)
    else:
        offer_data['categoryId'] = ''

    if offer.find('vendor') is not None:
        offer_data['vendor'] = offer.find('vendor').text
    else:
        offer_data['vendor'] = ""

    if offer.find('url') is not None:
        offer_data['url'] = offer.find('url').text
    else:
        offer_data['url'] = ""

    if offer.find('picture') is not None:
        offer_data['pictures'] = [picture.text for picture in offer.findall('picture')]
    else:
        offer_data['picture'] = ""

    if offer.find('description') is not None:
        offer_data['description'] = offer.find('description').text
    else:
        offer_data['description'] = ""

    data[vendor_code] = offer_data

print(data)
print(len(data))

# створюємо нову книгу Excel
workbook = openpyxl.Workbook()

# вибираємо активний лист
worksheet = workbook.active

# додаємо заголовки стовпців
worksheet['A1'] = 'A'
worksheet['B1'] = 'available'
worksheet['C1'] = 'price'
worksheet['D1'] = 'name'
worksheet['E1'] = 'categoryId'
worksheet['F1'] = 'vendor'
worksheet['G1'] = 'url'
worksheet['H1'] = 'pictures'
worksheet['I1'] = 'description'

# Отримуємо ключі словника
keys = list(data.keys())

# Додаємо дані в кожен рядок
for i in range(len(keys)):
    key = keys[i]

    row_num = i + 2  # рядок для запису даних, починаємо з другого рядка
    worksheet.cell(row=row_num, column=1, value=key)
    worksheet.cell(row=row_num, column=2, value=data[key]['available'])
    worksheet.cell(row=row_num, column=3, value=data[key]['price'])
    worksheet.cell(row=row_num, column=4, value=data[key]['name'])
    worksheet.cell(row=row_num, column=5, value=data[key]['categoryId'])
    worksheet.cell(row=row_num, column=6, value=data[key]['vendor'])
    worksheet.cell(row=row_num, column=7, value=data[key]['url'])
    worksheet.cell(row=row_num, column=8, value='; '.join(data[key]['pictures']))
    worksheet.cell(row=row_num, column=9, value=data[key]['description'])

# Зберігаємо книгу Excel
workbook.save("output.xlsx")

